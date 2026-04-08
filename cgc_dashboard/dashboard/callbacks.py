"""All Dash callbacks for the CGC dashboard."""

from __future__ import annotations
import json
import os
import io
import datetime
from typing import List

from dash import Input, Output, State, callback, html, no_update, ctx, ALL, dcc
import dash_bootstrap_components as dbc
import plotly.graph_objects as go

from cgc_dashboard.engine.constants import COMPONENTS
from cgc_dashboard.engine.cgc_stage import StageInputs, StageResult, calculate_stage
from cgc_dashboard.engine.flash_drum import FlashResult, run_flash
from cgc_dashboard.engine.train_runner import TrainConfig, TrainResult, run_train
from cgc_dashboard.dashboard.components.flash_panel import build_flash_content
from cgc_dashboard.dashboard.components.charts import (
    efficiency_chart, pressure_temperature_chart,
    condensate_chart, fouling_chart,
)
from cgc_dashboard.dashboard.components.stage_panel import stage_accordion_item
from cgc_dashboard.dashboard.components.train_overview import train_overview_strip


MAX_STAGES = 6


def _get_default_inputs():
    cfg_path = os.path.join(os.path.dirname(__file__), "..", "config", "default_inputs.json")
    with open(cfg_path, "r") as f:
        return json.load(f)


def _empty_fig():
    fig = go.Figure()
    fig.update_layout(
        template="plotly_dark", paper_bgcolor="#0a0a23",
        plot_bgcolor="#0a0a23", height=320,
        xaxis=dict(visible=False), yaxis=dict(visible=False),
        annotations=[dict(text="Run calculation to see results",
                          showarrow=False, font=dict(color="#555", size=14),
                          xref="paper", yref="paper", x=0.5, y=0.5)],
    )
    return fig


def register_callbacks(app):
    """Register all callbacks on the Dash app."""

    # ── LOAD DEFAULTS ───────────────────────────────────────────
    @app.callback(
        [Output(f"T-suc-{i}", "value") for i in range(MAX_STAGES)]
        + [Output(f"P-suc-{i}", "value") for i in range(MAX_STAGES)]
        + [Output(f"T-dis-{i}", "value") for i in range(MAX_STAGES)]
        + [Output(f"P-dis-{i}", "value") for i in range(MAX_STAGES)]
        + [Output(f"T-ac-{i}", "value") for i in range(MAX_STAGES)]
        + [Output(f"water-mf-{i}", "value") for i in range(MAX_STAGES)]
        + [Output(f"bfw-flow-{i}", "value") for i in range(MAX_STAGES)]
        + [Output(f"wo-flow-{i}", "value") for i in range(MAX_STAGES)]
        + [Output(f"wo-molar-{i}", "value") for i in range(MAX_STAGES)]
        + [Output(f"prev-dry-{i}", "value") for i in range(MAX_STAGES)]
        + [Output(f"prev-wmf-{i}", "value") for i in range(MAX_STAGES)]
        + [Output(f"prev-bfw-{i}", "value") for i in range(MAX_STAGES)]
        + [Output(f"prev-wo-{i}", "value") for i in range(MAX_STAGES)]
        + [Output(f"next-dry-{i}", "value") for i in range(MAX_STAGES)]
        + [Output(f"next-suc-p-{i}", "value") for i in range(MAX_STAGES)]
        + [Output(f"next-suc-t-{i}", "value") for i in range(MAX_STAGES)]
        + [Output(f"as1-active-{i}", "value") for i in range(MAX_STAGES)]
        + [Output(f"as1-analyser-{i}", "value") for i in range(MAX_STAGES)]
        + [Output(f"as1-feed-{i}", "value") for i in range(MAX_STAGES)]
        + [Output(f"as1-bottom-{i}", "value") for i in range(MAX_STAGES)]
        + [Output(f"as1-c3-{i}", "value") for i in range(MAX_STAGES)]
        + [Output(f"flash-dp-{i}", "value") for i in range(MAX_STAGES - 1)],
        Input("load-defaults-btn", "n_clicks"),
        State("current-num-stages", "data"),
        prevent_initial_call=True,
    )
    def load_defaults(n_clicks, num_stages):
        if not n_clicks:
            return [no_update] * (21 * MAX_STAGES + MAX_STAGES - 1)

        cfg = _get_default_inputs()
        num_stages = int(num_stages or 4)
        stages = cfg.get("stages", [])

        # Field extraction helpers
        def _val(stage_idx, key, default=0.0):
            if stage_idx < len(stages):
                return stages[stage_idx].get(key, default)
            return default

        outputs = []
        # T_suc, P_suc, T_dis, P_dis, T_ac, water_mf, bfw, wo, wo_molar
        for field_key in ["T_suc", "P_suc", "T_dis", "P_dis", "T_AC_out",
                          "water_mass_frac", "bfw_flow", "wash_oil_flow",
                          "wash_oil_molar_flow"]:
            for i in range(MAX_STAGES):
                outputs.append(_val(i, field_key))

        # prev_dry, prev_wmf, prev_bfw, prev_wo
        for field_key in ["prev_stage_dry_flow", "prev_water_mf",
                          "prev_bfw_flow", "prev_wash_oil_flow"]:
            for i in range(MAX_STAGES):
                outputs.append(_val(i, field_key))

        # next_dry, next_suc_p, next_suc_t
        for field_key in ["next_stage_dry_flow", "next_suc_pressure",
                          "next_suc_temperature"]:
            for i in range(MAX_STAGES):
                outputs.append(_val(i, field_key))

        # AS1 toggles and values
        for i in range(MAX_STAGES):
            outputs.append([1] if _val(i, "as1_active", False) else [])
        for i in range(MAX_STAGES):
            outputs.append([1] if _val(i, "as1_use_analyser", True) else [])
        for field_key in ["as1_feed_tph", "as1_bottom_tph", "as1_c3_pct"]:
            for i in range(MAX_STAGES):
                outputs.append(_val(i, field_key))

        # Flash dP (MAX_STAGES - 1 values)
        flash_dps = cfg.get("flash_dp", [0.3] * (MAX_STAGES - 1))
        for i in range(MAX_STAGES - 1):
            outputs.append(flash_dps[i] if i < len(flash_dps) else 0.3)

        return outputs

    # ── LOAD QO MASS FRACTIONS (pattern-matching callback) ──────
    @app.callback(
        Output({"type": "qo-mf", "stage": ALL, "comp": ALL}, "value"),
        Input("load-defaults-btn", "n_clicks"),
        State({"type": "qo-mf", "stage": ALL, "comp": ALL}, "id"),
        prevent_initial_call=True,
    )
    def load_qo_defaults(n_clicks, ids):
        if not n_clicks:
            return [no_update] * len(ids)

        cfg = _get_default_inputs()
        stages = cfg.get("stages", [])

        outputs = []
        for id_dict in ids:
            stage_idx = id_dict["stage"]
            comp = id_dict["comp"]
            val = 0.0
            if stage_idx < len(stages):
                val = stages[stage_idx].get(f"qo_mf_{comp}", 0.0)
            outputs.append(val)
        return outputs

    # ── MAIN CALCULATION ────────────────────────────────────────
    @app.callback(
        [Output("calc-results-store", "data"),
         Output("calc-timestamp", "children"),
         Output("error-banner", "children"),
         Output("error-banner", "style")]
        # KPI outputs per stage
        + [Output(f"kpi-pratio-{i}", "children") for i in range(MAX_STAGES)]
        + [Output(f"kpi-eff-{i}", "children") for i in range(MAX_STAGES)]
        + [Output(f"kpi-tdis-{i}", "children") for i in range(MAX_STAGES)]
        + [Output(f"kpi-head-{i}", "children") for i in range(MAX_STAGES)]
        + [Output(f"kpi-power-{i}", "children") for i in range(MAX_STAGES)]
        + [Output(f"kpi-mw-{i}", "children") for i in range(MAX_STAGES)]
        + [Output(f"kpi-acduty-{i}", "children") for i in range(MAX_STAGES)]
        + [Output(f"kpi-fouling-{i}", "children") for i in range(MAX_STAGES)]
        + [Output(f"kpi-hotapp-{i}", "children") for i in range(MAX_STAGES)]
        + [Output(f"kpi-ua-{i}", "children") for i in range(MAX_STAGES)]
        + [Output(f"kpi-volsuc-{i}", "children") for i in range(MAX_STAGES)]
        + [Output(f"kpi-voldis-{i}", "children") for i in range(MAX_STAGES)]
        + [Output(f"kpi-acdp-{i}", "children") for i in range(MAX_STAGES)]
        + [Output(f"kpi-normdp-{i}", "children") for i in range(MAX_STAGES)]
        + [Output(f"kpi-fouldp-{i}", "children") for i in range(MAX_STAGES)]
        + [Output(f"kpi-mwnobfw-{i}", "children") for i in range(MAX_STAGES)]
        # Overview strip
        + [Output(f"overview-eff-{i}", "children") for i in range(MAX_STAGES)]
        + [Output(f"overview-power-{i}", "children") for i in range(MAX_STAGES)]
        + [Output(f"overview-pratio-{i}", "children") for i in range(MAX_STAGES)]
        + [Output(f"overview-box-{i}", "style") for i in range(MAX_STAGES)]
        # Overview flash
        + [Output(f"overview-vf-{i}", "children") for i in range(MAX_STAGES - 1)]
        + [Output(f"overview-liq-{i}", "children") for i in range(MAX_STAGES - 1)]
        # Flash content
        + [Output(f"flash-content-{i}", "children") for i in range(MAX_STAGES - 1)]
        # Charts
        + [Output("chart-efficiency", "figure"),
           Output("chart-pt-profile", "figure"),
           Output("chart-condensate", "figure"),
           Output("chart-fouling", "figure")],
        Input("calc-btn", "n_clicks"),
        # Collect all stage inputs
        [State(f"T-suc-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"P-suc-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"T-dis-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"P-dis-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"T-ac-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"water-mf-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"bfw-flow-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"wo-flow-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"wo-molar-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"prev-dry-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"prev-wmf-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"prev-bfw-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"prev-wo-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"next-dry-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"next-suc-p-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"next-suc-t-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"as1-active-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"as1-analyser-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"as1-feed-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"as1-bottom-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"as1-c3-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"flash-dp-{i}", "value") for i in range(MAX_STAGES - 1)]
        + [State(f"auto-fill-flash-{i}", "value") for i in range(MAX_STAGES)]
        + [State({"type": "qo-mf", "stage": ALL, "comp": ALL}, "value")]
        + [State({"type": "qo-mf", "stage": ALL, "comp": ALL}, "id")]
        + [State("current-num-stages", "data")]
        + [State("dtl-path-input", "value")],
        prevent_initial_call=True,
    )
    def run_calculation(n_clicks, *args):
        if not n_clicks:
            total_outputs = 4 + 16 * MAX_STAGES + 4 * MAX_STAGES + 2 * (MAX_STAGES - 1) + (MAX_STAGES - 1) + 4
            return [no_update] * total_outputs

        # Parse state args
        idx = 0
        def _take(n):
            nonlocal idx
            vals = list(args[idx:idx+n])
            idx += n
            return vals

        t_suc = _take(MAX_STAGES)
        p_suc = _take(MAX_STAGES)
        t_dis = _take(MAX_STAGES)
        p_dis = _take(MAX_STAGES)
        t_ac = _take(MAX_STAGES)
        w_mf = _take(MAX_STAGES)
        bfw_f = _take(MAX_STAGES)
        wo_f = _take(MAX_STAGES)
        wo_m = _take(MAX_STAGES)
        prev_d = _take(MAX_STAGES)
        prev_w = _take(MAX_STAGES)
        prev_b = _take(MAX_STAGES)
        prev_o = _take(MAX_STAGES)
        next_d = _take(MAX_STAGES)
        next_p = _take(MAX_STAGES)
        next_t = _take(MAX_STAGES)
        as1_act = _take(MAX_STAGES)
        as1_ana = _take(MAX_STAGES)
        as1_feed = _take(MAX_STAGES)
        as1_bot = _take(MAX_STAGES)
        as1_c3 = _take(MAX_STAGES)
        flash_dp = _take(MAX_STAGES - 1)
        auto_fill = _take(MAX_STAGES)
        qo_vals = args[idx]; idx += 1
        qo_ids = args[idx]; idx += 1
        num_stages = int(args[idx] or 4); idx += 1
        dtl_path = args[idx] or ""; idx += 1

        # Build QO mass fraction dicts per stage
        qo_mf_dicts = [{} for _ in range(MAX_STAGES)]
        for id_dict, val in zip(qo_ids, qo_vals):
            s = id_dict["stage"]
            c = id_dict["comp"]
            qo_mf_dicts[s][c] = float(val or 0.0)

        def _f(v, default=0.0):
            try:
                return float(v) if v is not None else default
            except (TypeError, ValueError):
                return default

        # Build stage configs
        stage_cfgs = []
        for i in range(num_stages):
            cfg = {
                "T_suc": _f(t_suc[i]), "P_suc": _f(p_suc[i]),
                "T_dis": _f(t_dis[i]), "P_dis": _f(p_dis[i]),
                "T_AC_out": _f(t_ac[i]), "water_mass_frac": _f(w_mf[i]),
                "bfw_flow": _f(bfw_f[i]), "wash_oil_flow": _f(wo_f[i]),
                "wash_oil_molar_flow": _f(wo_m[i]),
                "prev_stage_dry_flow": _f(prev_d[i]),
                "prev_water_mf": _f(prev_w[i]),
                "prev_bfw_flow": _f(prev_b[i]),
                "prev_wash_oil_flow": _f(prev_o[i]),
                "next_stage_dry_flow": _f(next_d[i]),
                "next_suc_pressure": _f(next_p[i]),
                "next_suc_temperature": _f(next_t[i]),
                "as1_active": bool(as1_act[i]),
                "as1_use_analyser": bool(as1_ana[i]),
                "as1_feed_tph": _f(as1_feed[i]),
                "as1_bottom_tph": _f(as1_bot[i]),
                "as1_c3_pct": _f(as1_c3[i]),
            }
            for c in COMPONENTS:
                cfg[f"qo_mf_{c}"] = qo_mf_dicts[i].get(c, 0.0)
            stage_cfgs.append(cfg)

        # Build train config
        cfg_defaults = _get_default_inputs()
        train_cfg = TrainConfig(
            num_stages=num_stages,
            stages=stage_cfgs,
            global_params=cfg_defaults.get("global_params", {}),
            flash_dp=[_f(flash_dp[i], 0.3) for i in range(MAX_STAGES - 1)],
            dtl_path=dtl_path,
            auto_fill_from_flash=[bool(auto_fill[i]) for i in range(MAX_STAGES)],
        )

        # Run calculation
        train_result = run_train(train_cfg)

        # Build outputs
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        timestamp = f"Last calc: {now}"

        errors = train_result.errors
        error_banner = ""
        error_style = {"display": "none"}
        if errors:
            error_banner = dbc.Alert(
                [html.Strong("Errors: ")] + [html.Div(e) for e in errors],
                color="danger", dismissable=True,
            )
            error_style = {"display": "block", "marginTop": "8px"}

        # Store serializable results
        store_data = {
            "num_stages": num_stages,
            "timestamp": now,
        }

        # KPI outputs (16 fields × MAX_STAGES)
        kpi_fields = [
            ("P_ratio", ".4f"), ("efficiency", ".2f"), ("T_dis_BFW", ".2f"),
            ("polytropic_head", ".1f"), ("power", ".1f"), ("MW_withBFW", ".3f"),
            ("ac_duty", ".1f"), ("fouling_index", ".3f"), ("hot_approach", ".3f"),
            ("UA", ".1f"), ("vol_flow_suc", ".0f"), ("vol_flow_dis", ".0f"),
            ("ac_pressure_drop", ".3f"), ("norm_ac_dp", ".3f"),
            ("fouling_dp_pct", ".2f"), ("MW_noBFW", ".3f"),
        ]

        kpi_outputs = []
        for field_name, fmt in kpi_fields:
            for i in range(MAX_STAGES):
                if i < num_stages and i < len(train_result.stage_results):
                    val = getattr(train_result.stage_results[i], field_name, 0.0)
                    kpi_outputs.append(f"{val:{fmt}}")
                else:
                    kpi_outputs.append("—")

        # Overview strip KPIs
        overview_effs = []
        overview_powers = []
        overview_pratios = []
        overview_styles = []
        base_style = {
            "backgroundColor": "#16213e", "border": "2px solid #0f3460",
            "borderRadius": "8px", "padding": "8px 14px",
            "minWidth": "120px", "textAlign": "center",
        }
        for i in range(MAX_STAGES):
            if i < num_stages and i < len(train_result.stage_results):
                r = train_result.stage_results[i]
                overview_effs.append(f"Eff: {r.efficiency:.1f}%")
                overview_powers.append(f"Pwr: {r.power:.0f} kW")
                overview_pratios.append(f"PR: {r.P_ratio:.3f}")
                color = "#ff4444" if r.efficiency < 72 else "#ffa500" if r.efficiency < 78 else "#00ff88"
                overview_styles.append({**base_style, "borderColor": color})
            else:
                overview_effs.append("Eff: —")
                overview_powers.append("Pwr: —")
                overview_pratios.append("PR: —")
                overview_styles.append(base_style)

        # Flash overview
        overview_vfs = []
        overview_liqs = []
        flash_contents = []
        for i in range(MAX_STAGES - 1):
            if i < len(train_result.flash_results) and train_result.flash_results[i] is not None:
                f = train_result.flash_results[i]
                overview_vfs.append(f"V: {f.vapor_fraction:.4f}")
                overview_liqs.append(f"L: {f.liquid_flow_kmolh:.1f} kmol/h")
                flash_contents.append(build_flash_content(f, i))
            else:
                overview_vfs.append("V: —")
                overview_liqs.append("L: —")
                flash_contents.append(html.P("No flash result",
                                             style={"color": "#666"}))

        # Pad flash outputs if MAX_STAGES > num_stages
        while len(overview_vfs) < MAX_STAGES - 1:
            overview_vfs.append("V: —")
            overview_liqs.append("L: —")
            flash_contents.append(html.P("—", style={"color": "#666"}))

        # Charts
        sr = train_result.stage_results[:num_stages]
        si = train_result.stage_inputs_used[:num_stages]
        fr = train_result.flash_results[:num_stages]

        if sr:
            chart_eff = efficiency_chart(sr)
            chart_pt = pressure_temperature_chart(sr, si)
            chart_cond = condensate_chart(fr)
            chart_foul = fouling_chart(sr)
        else:
            chart_eff = chart_pt = chart_cond = chart_foul = _empty_fig()

        return (
            [store_data, timestamp, error_banner, error_style]
            + kpi_outputs
            + overview_effs + overview_powers + overview_pratios + overview_styles
            + overview_vfs + overview_liqs
            + flash_contents
            + [chart_eff, chart_pt, chart_cond, chart_foul]
        )

    # ── STAGE COUNT CHANGE ──────────────────────────────────────
    @app.callback(
        [Output("stages-accordion-container", "children"),
         Output("train-overview-container", "children"),
         Output("current-num-stages", "data")],
        Input("num-stages-select", "value"),
        prevent_initial_call=True,
    )
    def update_stage_count(value):
        n = int(value)
        accordion = dbc.Accordion(
            [stage_accordion_item(i, n) for i in range(n)],
            id="stages-accordion",
            start_collapsed=True, flush=True,
            style={"backgroundColor": "#0a0a23"},
        )
        overview = train_overview_strip(n)
        return accordion, overview, n

    # ── AS1 PANEL VISIBILITY ────────────────────────────────────
    for i in range(MAX_STAGES):
        @app.callback(
            Output(f"as1-panel-{i}", "style"),
            Input(f"as1-active-{i}", "value"),
            prevent_initial_call=True,
        )
        def toggle_as1(value, _i=i):
            base = {
                "backgroundColor": "#16213e", "border": "1px solid #0f3460",
                "borderRadius": "6px", "padding": "10px", "marginBottom": "8px",
            }
            if value:
                return base
            return {**base, "display": "none"}

    # ── EXCEL EXPORT ────────────────────────────────────────────
    @app.callback(
        Output("download-excel", "data"),
        Input("export-btn", "n_clicks"),
        State("calc-results-store", "data"),
        [State(f"T-suc-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"P-suc-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"T-dis-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"P-dis-{i}", "value") for i in range(MAX_STAGES)]
        + [State(f"T-ac-{i}", "value") for i in range(MAX_STAGES)]
        + [State("current-num-stages", "data")],
        prevent_initial_call=True,
    )
    def export_excel(n_clicks, store_data, *args):
        if not n_clicks or not store_data:
            return no_update

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            header_fill = PatternFill(start_color="1E1E1E", end_color="1E1E1E", fill_type="solid")
            header_font = Font(color="FFA500", bold=True, size=11)
            data_font = Font(color="FFFFFF", size=10)
            dark_fill1 = PatternFill(start_color="2A2A2A", end_color="2A2A2A", fill_type="solid")
            dark_fill2 = PatternFill(start_color="222222", end_color="222222", fill_type="solid")

            ws = wb.active
            ws.title = "Train Summary"
            ws.sheet_properties.tabColor = "FFA500"

            ws["A1"] = "CGC Train Performance Summary"
            ws["A1"].font = Font(color="FFA500", bold=True, size=14)
            ws["A2"] = f"Generated: {store_data.get('timestamp', '')}"
            ws["A2"].font = Font(color="999999", size=9)

            headers = ["KPI", "Unit"] + [f"Stage {i+1}" for i in range(store_data.get("num_stages", 4))]
            for j, h in enumerate(headers, 1):
                cell = ws.cell(row=4, column=j, value=h)
                cell.fill = header_fill
                cell.font = header_font

            ws["A5"] = "Exported from CGC Dashboard"
            ws["A5"].font = data_font

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return dcc.send_bytes(buf.getvalue(), "CGC_Train_Results.xlsx")

        except ImportError:
            return no_update
